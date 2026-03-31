from flask import Blueprint, render_template, redirect, url_for, send_file
from flask_login import login_required, current_user
from io import BytesIO
from datetime import datetime
import json

responses_bp = Blueprint('responses', __name__, url_prefix='/responses')

def serialize_responses(responses):
    """Convert MongoDB responses to JSON-safe dicts."""
    result = []
    for r in responses:
        clean = {}
        for k, v in r.items():
            if k == '_id':
                clean[k] = str(v)
            elif isinstance(v, datetime):
                clean[k] = v.isoformat()
            elif hasattr(v, '__str__') and not isinstance(v, (str, int, float, bool, list, dict, type(None))):
                clean[k] = str(v)
            else:
                clean[k] = v
        result.append(clean)
    return result

@responses_bp.route('/<form_id>')
@login_required
def view(form_id):
    from models.form import Form
    from models.response import Response
    form = Form.get_by_id(form_id)
    if not form or form['user_id'] != current_user.id:
        return redirect(url_for('dashboard.index'))
    raw_responses = Response.get_by_form(str(form['_id']))
    # Pre-serialize for template
    responses_json = json.dumps(serialize_responses(raw_responses))
    return render_template('responses/view.html',
                           form=form,
                           responses=raw_responses,
                           responses_json=responses_json,
                           form_id=form_id)

@responses_bp.route('/<form_id>/export')
@login_required
def export(form_id):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from models.form import Form
    from models.response import Response
    form = Form.get_by_id(form_id)
    if not form or form['user_id'] != current_user.id:
        return redirect(url_for('dashboard.index'))
    responses = Response.get_by_form(str(form['_id']))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Responses'
    fids, flabels = [], []
    for page in form.get('pages', []):
        for f in page.get('fields', []):
            fids.append(f['id'])
            flabels.append(f.get('label', f['id']))
    ws.append(['#', 'Submitted At'] + flabels)
    hfill = PatternFill(start_color='1A1A2E', end_color='1A1A2E', fill_type='solid')
    for cell in ws[1]:
        cell.fill = hfill
        cell.font = Font(color='FFB347', bold=True)
        cell.alignment = Alignment(horizontal='center')
    for i, resp in enumerate(responses, 1):
        ts = resp.get('submitted_at', '')
        ts_str = ts.strftime('%d/%m/%Y %H:%M') if isinstance(ts, datetime) else str(ts)
        row = [i, ts_str]
        for fid in fids:
            v = resp.get('data', {}).get(fid, '')
            row.append(', '.join(v) if isinstance(v, list) else str(v or ''))
        ws.append(row)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(c.value or '')) for c in col) + 4, 50)
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f"{form['title'].replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(out, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@responses_bp.route('/<form_id>/clear', methods=['POST'])
@login_required
def clear(form_id):
    from models.form import Form
    from models.response import Response
    form = Form.get_by_id(form_id)
    if form and form['user_id'] == current_user.id:
        Response.delete_all(str(form['_id']))
        Form.update(form_id, {'response_count': 0})
    return redirect(url_for('responses.view', form_id=form_id))
